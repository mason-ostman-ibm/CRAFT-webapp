#!/usr/bin/env python3
"""
Document Processor Service - Integrated from Mason's risk-document-completion
Provides RAG-powered Excel document completion with smart column detection
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, PatternFill
import openpyxl.utils
import os
import re
import warnings
import zipfile
import xml.etree.ElementTree as ET
from typing import List, Tuple
from ibm_watsonx_ai import Credentials
from ibm_watsonx_ai.foundation_models import ModelInference
from dotenv import load_dotenv
from astrapy import DataAPIClient
from sentence_transformers import SentenceTransformer
import logging

# Suppress benign openpyxl warnings about invalid specifications in Excel files
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

load_dotenv()
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize AstraDB for RAG
ASTRA_DB_API_ENDPOINT = os.getenv("ASTRA_DB_API_ENDPOINT")
ASTRA_DB_APPLICATION_TOKEN = os.getenv("ASTRA_DB_APPLICATION_TOKEN")

# Initialize clients
astra_client = None
collection = None
embedding_model = None

if ASTRA_DB_API_ENDPOINT and ASTRA_DB_APPLICATION_TOKEN:
    try:
        astra_client = DataAPIClient(ASTRA_DB_APPLICATION_TOKEN)
        astra_database = astra_client.get_database(ASTRA_DB_API_ENDPOINT)
        collection = astra_database.get_collection("qa_collection")
        embedding_model = SentenceTransformer('ibm-granite/granite-embedding-30m-english')
        logger.info("✓ RAG system initialized with AstraDB")
    except Exception as e:
        logger.warning(f"RAG system not available: {e}")
else:
    logger.warning("RAG system disabled - AstraDB credentials not configured")


def initialize_model():
    """Initialize the WatsonX.ai LLM model"""
    # Get credentials from environment
    watson_url = os.getenv("WATSON_URL")
    api_key = os.getenv("IBM_WATSONX_API_KEY")
    project_id = os.getenv("IBM_WATSONX_PROJECT_ID")
    space_id = os.getenv("SPACE_ID", "")
    username = os.getenv("CPD_USERNAME")
    model_id = os.getenv("WATSON_TEXT_MODEL")

    # Create credentials
    credentials = Credentials(
        url=watson_url,
        username=username,
        api_key=api_key
    )

    # Model parameters
    parameters = {
        "frequency_penalty": 0,
        "max_tokens": 2000,
        "presence_penalty": 0,
        "temperature": 0.7,
        "top_p": 1
    }

    # Initialize model
    model = ModelInference(
        model_id=model_id,
        params=parameters,
        credentials=credentials,
        project_id=project_id,
        space_id=space_id if space_id else None
    )

    logger.info(f"✓ Model initialized: {model_id}")
    return model


def detect_qa_columns_in_sheet(df, model):
    """
    Use LLM to intelligently detect which columns contain questions and answers

    Args:
        df: pandas DataFrame of the sheet
        model: Initialized LLM model

    Returns:
        tuple: (question_column_index, answer_column_index) or (None, None) - 1-based indices
    """
    # Get available column names
    column_names = list(df.columns)

    # If there are only 2 columns, just use them
    if len(column_names) == 2:
        logger.info(f"✓ Auto-detected 2 columns - Q: {column_names[0]} (col 1), A: {column_names[1]} (col 2)")
        return 1, 2

    sample_data = df.head(5).to_string()

    prompt = f"""Given this spreadsheet data, identify which column NUMBER (1, 2, 3, etc.) contains questions and which contains answers.

Available columns: {', '.join([f'{i+1}: {col}' for i, col in enumerate(column_names)])}

Sample data:
{sample_data}

Respond in this exact format (use only the column NUMBER):
Question column: [number]
Answer column: [number]"""

    messages = [
        {
            "role": "user",
            "content": prompt
        }
    ]

    try:
        response = model.chat(messages=messages)
        answer = response["choices"][0]["message"]["content"]

        # Parse the response - look for numbers
        question_match = re.search(r'Question column:\s*(\d+)', answer)
        answer_match = re.search(r'Answer column:\s*(\d+)', answer)

        if question_match and answer_match:
            question_col_idx = int(question_match.group(1))
            answer_col_idx = int(answer_match.group(1))

            # Validate indices
            if 1 <= question_col_idx <= len(column_names) and 1 <= answer_col_idx <= len(column_names):
                logger.info(f"✓ Detected columns - Q: {column_names[question_col_idx-1]} (col {question_col_idx}), A: {column_names[answer_col_idx-1]} (col {answer_col_idx})")
                return question_col_idx, answer_col_idx
            else:
                logger.error(f"Invalid column indices detected: Q={question_col_idx}, A={answer_col_idx}")
                return None, None

        return None, None
    except Exception as e:
        logger.error(f"Error detecting columns: {e}")
        return None, None


def get_relevant_context(question, top_k=5, similarity_threshold=0.5):
    """
    Search for relevant Q&A pairs using RAG with AstraDB
    
    Args:
        question: The user's question
        top_k: Number of top matches to return
        similarity_threshold: Minimum similarity score
    
    Returns:
        String formatted context for LLM prompt
    """
    if not collection or not embedding_model:
        return "No RAG context available. Use general knowledge."

    try:
        query_embedding = embedding_model.encode(question).tolist()

        results = collection.find(
            sort={"$vector": query_embedding},
            limit=top_k,
            projection={"question": 1, "answer": 1, "category": 1},
            include_similarity=True
        )

        context = ""
        relevant_count = 0

        for result in results:
            similarity = result.get('$similarity', 0)

            if similarity < similarity_threshold:
                continue

            q = result.get('question', 'N/A')
            a = result.get('answer', 'N/A')

            answer_lower = str(a).lower().strip()
            if answer_lower in ['unanswered', 'nan', 'none', '', 'n/a']:
                continue

            relevant_count += 1
            context += f"Example {relevant_count}:\n"
            context += f"Q: {q}\n"
            context += f"A: {a}\n\n"

        if not context:
            context = "No relevant examples found. Use your general knowledge about IBM and business practices."

        logger.info(f"✓ Retrieved {relevant_count} relevant examples for RAG")
        return context.strip()

    except Exception as e:
        logger.error(f"Error retrieving RAG context: {e}")
        return "RAG retrieval error. Use general knowledge."


def ask_llm(question, model, context=""):
    """
    Ask the LLM a question with optional RAG context
    
    Args:
        question: The question to answer
        model: Initialized LLM model
        context: Additional context (optional)
    
    Returns:
        Generated answer
    """
    try:
        # Get RAG context if available
        rag_context = get_relevant_context(question, top_k=5, similarity_threshold=0.5)
        
        # Combine with any additional context
        full_context = f"{rag_context}\n\n{context}" if context else rag_context

        messages = [
            {
                "role": "system",
                "content": """You are a professional document completion assistant for IBM. Your role is to fill out governance, compliance, and business documents with accurate, concise information.

INSTRUCTIONS:
1. You are filling out forms and documents on behalf of IBM
2. Answer questions directly and professionally, as if completing an official form
3. Use the provided context examples as reference for style, format, and content
4. Match the tone and detail level of the context examples
5. Be concise - forms require brief, direct answers, not explanations
6. If context is provided, adapt it to answer the specific question
7. Do NOT mention that you're using context or reference materials
8. Do NOT include meta-commentary like "based on the context"
9. NEVER respond with just "unanswered" or leave the answer blank
10. If you don't have relevant information, write "Information not available"

FORMATTING:
- For yes/no questions: Answer "Yes" or "No" followed by brief details if needed
- For descriptive questions: Provide 1-3 sentences maximum unless more detail is clearly needed
- For lists: Use bullet points or numbered lists as appropriate
- Maintain professional business language throughout

Remember: You ARE the person filling out this document. Write answers directly as they should appear in the form."""
            },
            {
                "role": "user",
                "content": f"""Using the following reference examples, answer the question below.

REFERENCE EXAMPLES:
{full_context}

QUESTION TO ANSWER:
{question}

Provide your answer:"""
            }
        ]

        response = model.chat(messages=messages)
        answer = response["choices"][0]["message"]["content"]

        if not answer or answer.strip() == "":
            return "Error: Empty response from model"

        return answer

    except Exception as e:
        logger.error(f"Error asking LLM: {e}")
        return f"Error: {str(e)}"


def should_skip_sheet(sheet_name):
    """
    Determine if a sheet should be skipped (instruction sheets, empty sheets, etc.)

    Args:
        sheet_name: Name of the worksheet

    Returns:
        True if sheet should be skipped, False otherwise
    """
    if not sheet_name:
        return True

    sheet_lower = sheet_name.lower()

    # Skip instruction/system sheets
    skip_keywords = ['instruction', 'dv_sheet', 'legend']

    return any(skip in sheet_lower for skip in skip_keywords)


def get_sheet_names_xml(file_path) -> List[str]:
    """
    Get sheet names from Excel file using XML parser
    Fallback method for files that openpyxl can't read properly

    Args:
        file_path: Path to Excel file

    Returns:
        List of sheet names
    """
    try:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            wb_xml = zip_ref.read('xl/workbook.xml').decode('utf-8')
            sheet_info = re.findall(r'<sheet.*?name="([^"]+)"', wb_xml)
            return sheet_info
    except Exception:
        return []


def get_sheet_qa_data_pandas(file_path, sheet_name, question_col_name, answer_col_name):
    """
    Read Q&A data from a specific sheet using pandas

    Args:
        file_path: Path to Excel file
        sheet_name: Name of the sheet to read
        question_col_name: Name of question column
        answer_col_name: Name of answer column

    Returns:
        List of tuples: [(row_num, question, has_answer), ...]
    """
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        if df.empty or question_col_name not in df.columns or answer_col_name not in df.columns:
            return []

        qa_pairs = []
        for idx, row in df.iterrows():
            question = row.get(question_col_name)
            answer = row.get(answer_col_name)

            question_str = str(question).strip() if pd.notna(question) else ""
            answer_str = str(answer).strip() if pd.notna(answer) else ""

            # Skip empty questions
            if not question_str or question_str == 'nan':
                continue

            # Check if answer exists
            has_answer = bool(answer_str and answer_str not in ['nan', 'unanswered', ''])

            # Excel row number is pandas index + 2 (header + 0-based indexing)
            qa_pairs.append((idx + 2, question_str, has_answer))

        return qa_pairs

    except Exception as e:
        logger.error(f"Error reading sheet with pandas: {e}")
        return []


def process_document(file_path, output_path=None, context=""):
    """
    Main function to process Excel document with RAG-powered answers
    
    Args:
        file_path: Path to input Excel file
        output_path: Path to output Excel file (optional)
        context: Additional context for AI processing (optional)
    
    Returns:
        dict: Processing results with statistics
    """
    try:
        logger.info(f"{'='*60}")
        logger.info("EXCEL AI PROCESSOR - Document Processing")
        logger.info(f"{'='*60}")

        # Validate input file exists
        if not os.path.exists(file_path):
            error_msg = f"Input file not found: {file_path}"
            logger.error(error_msg)
            return {
                "success": False,
                "error": error_msg,
                "output_path": None,
                "sheets_processed": 0,
                "total_sheets": 0,
                "questions_answered": 0,
                "details": []
            }

        # Initialize model
        logger.info("Initializing WatsonX model...")
        try:
            model = initialize_model()
        except Exception as e:
            error_msg = f"Failed to initialize model: {str(e)}"
            logger.error(error_msg)
            return {
                "success": False,
                "error": error_msg,
                "output_path": None,
                "sheets_processed": 0,
                "total_sheets": 0,
                "questions_answered": 0,
                "details": []
            }

        # Load workbook with warnings suppressed
        logger.info(f"Loading workbook: {file_path}")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=False)

            # Check if the file has corrupted sheet structure
            use_xml_fallback = len(wb.sheetnames) == 0

            if use_xml_fallback:
                logger.warning("File has corrupted sheet structure - using XML fallback")
                sheet_names = get_sheet_names_xml(file_path)
                if not sheet_names:
                    raise Exception("No sheets found in workbook. File may be corrupted or empty.")
            else:
                sheet_names = wb.sheetnames
                logger.info(f"Found {len(sheet_names)} sheets")

        except Exception as e:
            error_msg = f"Failed to load workbook: {str(e)}"
            logger.error(error_msg)
            return {
                "success": False,
                "error": error_msg,
                "output_path": None,
                "sheets_processed": 0,
                "total_sheets": 0,
                "questions_answered": 0,
                "details": []
            }

        total_questions_answered = 0
        sheets_processed = 0
        processing_details = []
        qa_pairs = []
        errors = []

        # Process each sheet
        for sheet_name in sheet_names:
            try:
                # Check if we should skip this sheet
                if should_skip_sheet(sheet_name):
                    logger.info(f"⊘ Skipping sheet (instruction/legend): {sheet_name}")
                    continue

                logger.info(f"\n{'='*60}")
                logger.info(f"Processing sheet: {sheet_name}")
                logger.info(f"{'='*60}")

                # Get or create worksheet
                try:
                    ws = wb[sheet_name]
                except KeyError:
                    # Sheet exists in XML but not accessible - create new one
                    logger.warning(f"  ⚠ Creating new sheet '{sheet_name}' (original not accessible)")
                    ws = wb.create_sheet(sheet_name)

                # Read sheet with error handling
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                except Exception as e:
                    logger.error(f"  ✗ Failed to read sheet: {str(e)}")
                    errors.append(f"Sheet '{sheet_name}': Failed to read - {str(e)}")
                    continue

                if df.empty:
                    logger.info(f"  ⊘ Skipping empty sheet")
                    continue

                # Detect Q&A columns using LLM (returns 1-based indices)
                logger.info(f"  Detecting Q&A columns...")
                try:
                    question_col_idx, answer_col_idx = detect_qa_columns_in_sheet(df, model)
                except Exception as e:
                    logger.error(f"  ✗ Column detection failed: {str(e)}")
                    errors.append(f"Sheet '{sheet_name}': Column detection failed - {str(e)}")
                    continue

                if not question_col_idx or not answer_col_idx:
                    logger.info(f"  ⊘ Could not detect Q&A columns, skipping sheet")
                    errors.append(f"Sheet '{sheet_name}': Could not detect Q&A columns")
                    continue

                # Get actual column names for reference (indices are 1-based, pandas columns are 0-based)
                try:
                    question_col_name = df.columns[question_col_idx - 1]
                    answer_col_name = df.columns[answer_col_idx - 1]
                    logger.info(f"  ✓ Question column: {question_col_name} (index {question_col_idx})")
                    logger.info(f"  ✓ Answer column: {answer_col_name} (index {answer_col_idx})")
                except IndexError as e:
                    logger.error(f"  ✗ Invalid column indices: Q={question_col_idx}, A={answer_col_idx}")
                    errors.append(f"Sheet '{sheet_name}': Invalid column indices detected")
                    continue

                # Set column widths safely
                try:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(question_col_idx)].width = 60
                    ws.column_dimensions[openpyxl.utils.get_column_letter(answer_col_idx)].width = 80
                except Exception as e:
                    logger.warning(f"  ⚠ Could not set column widths: {str(e)}")

                # Process rows
                questions_in_sheet = 0
                for idx, row in df.iterrows():
                    try:
                        excel_row = idx + 2

                        # Access using column index (0-based for pandas)
                        try:
                            question = row.iloc[question_col_idx - 1]
                            answer = row.iloc[answer_col_idx - 1]
                        except IndexError as e:
                            logger.error(f"  ✗ Error accessing columns in row {excel_row}: {str(e)}")
                            continue

                        if pd.isna(question) or str(question).strip() == "":
                            continue

                        question_str = str(question).strip()
                        answer_str = str(answer).strip() if pd.notna(answer) else ""

                        # Check if answer is missing (consistent with delta_service logic)
                        answer_lower = answer_str.lower()
                        _unanswered_values = [
                            'nan', 'none', 'unanswered', 'tbd', 'to be determined',
                            'pending', 'todo', 'blank', 'empty', 'n/a'
                        ]
                        is_unanswered = (
                            not answer_str or
                            answer_str.isspace() or
                            answer_lower in _unanswered_values or
                            (len(answer_str) == 1 and answer_str in ['-', '_']) or
                            answer_str in ['--', '---']
                        )

                        if is_unanswered:
                            logger.info(f"  Answering row {excel_row}: {question_str[:60]}...")

                            # Get answer from LLM with RAG
                            try:
                                generated_answer = ask_llm(question_str, model, context)
                            except Exception as e:
                                logger.error(f"  ✗ Failed to generate answer for row {excel_row}: {str(e)}")
                                generated_answer = f"Error generating answer: {str(e)}"

                            # Handle merged cells safely
                            try:
                                answer_cell = ws.cell(row=excel_row, column=answer_col_idx)
                                if isinstance(answer_cell, openpyxl.cell.cell.MergedCell):
                                    for merged_range in list(ws.merged_cells.ranges):
                                        if answer_cell.coordinate in merged_range:
                                            ws.unmerge_cells(str(merged_range))
                                            break
                                    answer_cell = ws.cell(row=excel_row, column=answer_col_idx)

                                # Fill in the answer and highlight as AI-generated
                                answer_cell.value = generated_answer
                                answer_cell.alignment = Alignment(wrap_text=True, vertical='top')
                                answer_cell.fill = PatternFill(
                                    start_color="FFFACD",
                                    end_color="FFFACD",
                                    fill_type="solid"
                                )
                            except Exception as e:
                                logger.error(f"  ✗ Failed to write answer to row {excel_row}: {str(e)}")
                                continue

                            # Format question cell safely
                            try:
                                question_cell = ws.cell(row=excel_row, column=question_col_idx)
                                if isinstance(question_cell, openpyxl.cell.cell.MergedCell):
                                    for merged_range in list(ws.merged_cells.ranges):
                                        if question_cell.coordinate in merged_range:
                                            ws.unmerge_cells(str(merged_range))
                                            break
                                    question_cell = ws.cell(row=excel_row, column=question_col_idx)

                                question_cell.alignment = Alignment(wrap_text=True, vertical='top')
                            except Exception as e:
                                logger.warning(f"  ⚠ Could not format question cell at row {excel_row}: {str(e)}")

                            # Adjust row height safely
                            try:
                                estimated_lines = max(len(generated_answer) // 80, 1)
                                min_height = max(15 * estimated_lines, 15)
                                ws.row_dimensions[excel_row].height = min(min_height, 150)
                            except Exception as e:
                                logger.warning(f"  ⚠ Could not adjust row height for row {excel_row}: {str(e)}")

                            qa_pairs.append({
                                'sheet': sheet_name,
                                'row': excel_row,
                                'question': question_str,
                                'answer': generated_answer
                            })
                            questions_in_sheet += 1
                            total_questions_answered += 1

                    except Exception as e:
                        logger.error(f"  ✗ Error processing row {idx + 2}: {str(e)}")
                        errors.append(f"Sheet '{sheet_name}', Row {idx + 2}: {str(e)}")
                        continue

                sheets_processed += 1
                processing_details.append({
                    "sheet": sheet_name,
                    "questions_answered": questions_in_sheet
                })
                logger.info(f"  ✓ Answered {questions_in_sheet} questions in this sheet")

            except Exception as e:
                logger.error(f"  ✗ Unexpected error processing sheet '{sheet_name}': {str(e)}")
                errors.append(f"Sheet '{sheet_name}': Unexpected error - {str(e)}")
                continue

        # Save workbook
        if output_path is None:
            from pathlib import Path
            input_path = Path(file_path)
            output_path = str(input_path.parent / f"{input_path.stem}_completed{input_path.suffix}")

        logger.info(f"\n{'='*60}")
        logger.info("PROCESSING SUMMARY")
        logger.info(f"{'='*60}")
        logger.info(f"Sheets processed: {sheets_processed}/{len(sheet_names)}")
        logger.info(f"Total questions answered: {total_questions_answered}")
        logger.info(f"Output file: {output_path}")
        if errors:
            logger.warning(f"Errors encountered: {len(errors)}")
            for error in errors[:5]:  # Show first 5 errors
                logger.warning(f"  - {error}")
        logger.info(f"{'='*60}\n")

        # Save with error handling
        try:
            wb.save(output_path)
            logger.info(f"✓ File saved successfully to: {output_path}")
        except Exception as e:
            error_msg = f"Failed to save workbook: {str(e)}"
            logger.error(error_msg)
            return {
                "success": False,
                "error": error_msg,
                "output_path": None,
                "sheets_processed": sheets_processed,
                "total_sheets": len(sheet_names),
                "questions_answered": total_questions_answered,
                "details": processing_details,
                "errors": errors
            }

        return {
            "success": True,
            "output_path": output_path,
            "sheets_processed": sheets_processed,
            "total_sheets": len(sheet_names),
            "questions_answered": total_questions_answered,
            "details": processing_details,
            "qa_pairs": qa_pairs,
            "errors": errors if errors else None
        }

    except Exception as e:
        error_msg = f"Unexpected error in process_document: {str(e)}"
        logger.error(error_msg)
        import traceback
        logger.error(traceback.format_exc())
        return {
            "success": False,
            "error": error_msg,
            "output_path": None,
            "sheets_processed": 0,
            "total_sheets": 0,
            "questions_answered": 0,
            "details": [],
            "errors": [error_msg]
        }


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python document_processor.py <input_file.xlsx> [output_file.xlsx] [context]")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    context = sys.argv[3] if len(sys.argv) > 3 else ""

    result = process_document(input_file, output_file, context)
    print(f"\n✓ Processing complete!")
    print(f"Questions answered: {result['questions_answered']}")
    print(f"File saved to: {result['output_path']}")

# Made with Bob
