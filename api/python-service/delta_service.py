#!/usr/bin/env python3
"""
Delta Service - Questionnaire Delta Tool
Provides intelligent answer reuse from previous year questionnaires
using vector similarity matching with AstraDB and IBM Granite embeddings.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
import openpyxl.utils
import os
import logging
from datetime import datetime
from pathlib import Path
from astrapy import DataAPIClient
from astrapy.info import CollectionDefinition, CollectionVectorOptions
from sentence_transformers import SentenceTransformer
from ibm_watsonx_ai import Credentials
from ibm_watsonx_ai.foundation_models import ModelInference
from dotenv import load_dotenv
from document_processor import detect_qa_columns_in_sheet, should_skip_sheet

load_dotenv()
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configuration
ASTRA_DB_API_ENDPOINT = os.getenv("ASTRA_DB_API_ENDPOINT")
ASTRA_DB_APPLICATION_TOKEN = os.getenv("ASTRA_DB_APPLICATION_TOKEN")
HIGH_CONFIDENCE_THRESHOLD = 0.90
MEDIUM_CONFIDENCE_THRESHOLD = float(os.getenv("DELTA_SIMILARITY_THRESHOLD", "0.85"))
ENABLE_LLM_VERIFICATION = os.getenv("DELTA_ENABLE_LLM_VERIFICATION", "true").lower() == "true"
USE_LLM_GENERATION = os.getenv("DELTA_USE_LLM_GENERATION", "false").lower() == "true"

class DeltaService:
    """Delta Service for intelligent questionnaire answer reuse"""
    
    def __init__(self):
        self.astra_client = None
        self.db = None
        self.embedding_model = None
        self.watsonx_model = None
        self.embedding_cache = {}
        
        # Initialize AstraDB
        if ASTRA_DB_API_ENDPOINT and ASTRA_DB_APPLICATION_TOKEN:
            try:
                self.astra_client = DataAPIClient(ASTRA_DB_APPLICATION_TOKEN)
                self.db = self.astra_client.get_database_by_api_endpoint(ASTRA_DB_API_ENDPOINT)
                logger.info("✓ AstraDB initialized")
            except Exception as e:
                logger.warning(f"AstraDB initialization failed: {e}")
        else:
            logger.warning("AstraDB credentials not configured")
        
        # Initialize embedding model
        try:
            self.embedding_model = SentenceTransformer('ibm-granite/granite-embedding-30m-english')
            logger.info("✓ Embedding model initialized")
        except Exception as e:
            logger.warning(f"Embedding model initialization failed: {e}")
    
    def is_available(self):
        """Check if Delta Service is available"""
        return self.db is not None and self.embedding_model is not None
    
    def initialize_watsonx_model(self):
        """Initialize WatsonX model for LLM verification"""
        if self.watsonx_model is not None:
            return self.watsonx_model

        try:
            credentials = Credentials(
                url=os.getenv("WATSON_URL", "https://ca-tor.ml.cloud.ibm.com"),
                username=os.getenv("CPD_USERNAME"),
                api_key=os.getenv("IBM_WATSONX_API_KEY")
            )

            project_id = os.getenv("IBM_WATSONX_PROJECT_ID")
            model_id = os.getenv("WATSON_TEXT_MODEL", "meta-llama/llama-3-3-70b-instruct")
            

            parameters = {
                "temperature": 0.1,
                "max_tokens": 100,
                "top_p": 1
            }

            self.watsonx_model = ModelInference(
                model_id=model_id,
                params=parameters,
                credentials=credentials,
                project_id=project_id
            )

            logger.info(f"✓ WatsonX model initialized: {model_id}")
            return self.watsonx_model
        except Exception as e:
            logger.error(f"Failed to initialize WatsonX model: {e}")
            return None
    
    def generate_embedding(self, text):
        """Generate embedding for text using IBM Granite model"""
        # Check cache
        if text in self.embedding_cache:
            return self.embedding_cache[text]
        
        try:
            embedding = self.embedding_model.encode(text).tolist()
            self.embedding_cache[text] = embedding
            return embedding
        except Exception as e:
            logger.error(f"Error generating embedding: {e}")
            raise
    
    def extract_qa_pairs(self, file_path):
        """Extract Q&A pairs from Excel file using smart column detection"""
        logger.info("=" * 60)
        logger.info("EXTRACT_QA_PAIRS - NEW CODE VERSION WITH NA FIX")
        logger.info("=" * 60)
        qa_data = []

        # Initialize WatsonX model for column detection
        model = self.initialize_watsonx_model()
        if model is None:
            logger.error("Cannot extract Q&A pairs without WatsonX model for column detection")
            raise Exception("WatsonX model required for smart column detection")

        # Load workbook
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in workbook.sheetnames:
            # Skip instruction/legend sheets
            if should_skip_sheet(sheet_name):
                logger.info(f"Skipping sheet: {sheet_name}")
                continue

            try:
                # Read sheet as DataFrame for column detection
                df = pd.read_excel(file_path, sheet_name=sheet_name)

                if df.empty:
                    logger.info(f"Skipping empty sheet: {sheet_name}")
                    continue

                # Use smart column detection (returns 1-based indices)
                question_col_idx, answer_col_idx = detect_qa_columns_in_sheet(df, model)

                if not question_col_idx or not answer_col_idx:
                    logger.warning(f"Could not detect Q&A columns in sheet: {sheet_name}")
                    continue

                # Extract Q&A pairs using detected columns
                for idx, row in df.iterrows():
                    try:
                        # Access using detected column indices (convert to 0-based for pandas)
                        question = row.iloc[question_col_idx - 1]
                        answer = row.iloc[answer_col_idx - 1]

                        # Skip if question is not valid
                        if pd.isna(question):
                            continue

                        question_str = str(question).strip()
                        
                        # Skip empty questions
                        if not question_str:
                            continue
                        
                        # Handle answer - convert NaN to empty string
                        answer_str = str(answer).strip() if pd.notna(answer) else ''

                        # Debug: Log the raw answer value for troubleshooting
                        if answer_str.upper() in ['NA', 'N/A']:
                            logger.info(f"Row {idx + 2}: Raw answer='{answer_str}' (repr={repr(answer_str)}), len={len(answer_str)}")

                        # Check if answer is actually filled in
                        # Be more specific - only treat truly empty or explicitly unanswered values as unanswered
                        # "NA" by itself could mean "Not Applicable" which IS an answer
                        unanswered_values = [
                            'unanswered',
                            'tbd', 'to be determined',
                            'pending',
                            'todo',
                            'blank',
                            'empty'
                        ]

                        answer_lower = answer_str.lower().strip()
                        
                        # Check each condition separately for debugging
                        is_empty = not answer_str
                        is_whitespace = answer_str.isspace() if answer_str else False
                        is_in_unanswered_list = answer_lower in unanswered_values
                        is_single_dash = len(answer_str) == 1 and answer_str in ['-', '_']
                        is_multiple_dash = answer_str in ['--', '---']
                        
                        is_unanswered = (
                            is_empty or
                            is_whitespace or
                            is_in_unanswered_list or
                            is_single_dash or
                            is_multiple_dash
                        )

                        # Include ALL questions - both answered and unanswered
                        # For unanswered questions, store "UNANSWERED" as the answer
                        final_answer = 'UNANSWERED' if is_unanswered else answer_str
                        
                        # Debug logging for "NA" answers
                        if answer_str.upper() in ['NA', 'N/A']:
                            logger.info(f"Row {idx + 2}: Checks - empty={is_empty}, whitespace={is_whitespace}, in_list={is_in_unanswered_list}, dash={is_single_dash or is_multiple_dash}")
                            logger.info(f"Row {idx + 2}: answer_lower='{answer_lower}', is_unanswered={is_unanswered}, final_answer='{final_answer}'")
                        
                        qa_data.append({
                            'question': question_str,
                            'answer': final_answer,
                            'sheet': sheet_name,
                            'row': idx + 2,  # Excel row (header + 0-based index)
                            'question_col': question_col_idx - 1,  # Store 0-based for consistency
                            'answer_col': answer_col_idx - 1
                        })
                        
                        if is_unanswered and answer_str.upper() not in ['NA', 'N/A']:
                            logger.debug(f"Including unanswered question at row {idx + 2}: {question_str[:60]}...")
                    except Exception as e:
                        logger.warning(f"Error extracting row {idx + 2} in sheet {sheet_name}: {e}")
                        continue

            except Exception as e:
                logger.error(f"Error processing sheet {sheet_name}: {e}")
                continue

        return qa_data
    
    def extract_questions(self, file_path):
        """Extract questions from Excel file using smart column detection"""
        questions = []

        # Initialize WatsonX model for column detection
        model = self.initialize_watsonx_model()
        if model is None:
            logger.error("Cannot extract questions without WatsonX model for column detection")
            raise Exception("WatsonX model required for smart column detection")

        # Load workbook
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in workbook.sheetnames:
            # Skip instruction/legend sheets
            if should_skip_sheet(sheet_name):
                logger.info(f"Skipping sheet: {sheet_name}")
                continue

            try:
                # Read sheet as DataFrame for column detection
                df = pd.read_excel(file_path, sheet_name=sheet_name)

                if df.empty:
                    logger.info(f"Skipping empty sheet: {sheet_name}")
                    continue

                # Use smart column detection (returns 1-based indices)
                question_col_idx, answer_col_idx = detect_qa_columns_in_sheet(df, model)

                if not question_col_idx or not answer_col_idx:
                    logger.warning(f"Could not detect Q&A columns in sheet: {sheet_name}")
                    continue

                # Extract questions using detected columns
                for idx, row in df.iterrows():
                    try:
                        # Access using detected column indices (convert to 0-based for pandas)
                        question = row.iloc[question_col_idx - 1]

                        if pd.isna(question):
                            continue

                        question_str = str(question).strip()

                        if question_str:
                            # Get current answer if exists
                            current_answer = ''
                            try:
                                answer = row.iloc[answer_col_idx - 1]
                                if pd.notna(answer):
                                    current_answer = str(answer).strip()
                            except:
                                pass

                            questions.append({
                                'question': question_str,
                                'sheet': sheet_name,
                                'row': idx + 2,  # Excel row (header + 0-based index)
                                'question_col': question_col_idx - 1,  # Store 0-based
                                'answer_col': answer_col_idx - 1,
                                'current_answer': current_answer
                            })
                    except Exception as e:
                        logger.warning(f"Error extracting row {idx + 2} in sheet {sheet_name}: {e}")
                        continue

            except Exception as e:
                logger.error(f"Error processing sheet {sheet_name}: {e}")
                continue

        return questions
    
    def ingest_baseline(self, file_path, user_id, description=''):
        """Ingest baseline questionnaire into AstraDB (clears existing baseline)"""
        if not self.is_available():
            raise Exception('Delta Service not available - AstraDB or embedding model not configured')

        logger.info(f"\n{'='*60}")
        logger.info('DELTA SERVICE - Baseline Ingestion')
        logger.info(f"{'='*60}")
        logger.info(f"File: {file_path}")
        logger.info(f"User: {user_id}")

        # Extract Q&A pairs using smart column detection
        qa_data = self.extract_qa_pairs(file_path)

        logger.info(f"Extracted {len(qa_data)} Q&A pairs")

        if len(qa_data) == 0:
            raise Exception('No valid Q&A pairs found in the file')

        # Use single delta_baseline collection
        collection_name = "delta_baseline"

        # Check if collection exists, create or clear it
        try:
            existing_collections = self.db.list_collection_names()
            logger.info(f"Existing collections: {existing_collections}")

            if collection_name in existing_collections:
                logger.info(f"Collection exists - clearing all existing data: {collection_name}")
                collection = self.db.get_collection(collection_name)
                # Delete all documents in the collection
                delete_result = collection.delete_many({})
                logger.info(f"✓ Cleared existing baseline data (deleted documents)")
            else:
                logger.info(f"Creating new collection: {collection_name}")
                collection = self.db.create_collection(
                    name=collection_name,
                    definition=CollectionDefinition(
                        vector=CollectionVectorOptions(
                            dimension=384,
                            metric="cosine"
                        )
                    )
                )
                logger.info(f"✓ Created new collection: {collection_name}")
        except Exception as e:
            logger.error(f"Error with collection management: {e}")
            logger.error(f"Error type: {type(e).__name__}")
            import traceback
            logger.error(traceback.format_exc())
            raise Exception(f"Failed to create/get collection {collection_name}: {str(e)}")
        
        # Generate embeddings and insert documents
        logger.info('Generating embeddings...')
        documents = []
        
        for i, qa in enumerate(qa_data):
            embedding = self.generate_embedding(qa['question'])

            documents.append({
                'question_text': qa['question'],
                'answer_text': qa['answer'],
                '$vector': embedding,
                'metadata': {
                    'sheet_name': qa['sheet'],
                    'row_number': qa['row'],
                    'question_col': qa['question_col'],
                    'answer_col': qa['answer_col'],
                    'user_email': user_id,
                    'description': description,
                    'ingested_at': datetime.utcnow().isoformat()
                }
            })
            
            if (i + 1) % 10 == 0:
                logger.info(f"  Progress: {i + 1}/{len(qa_data)} embeddings generated")
        
        # Verify collection exists before inserting
        try:
            existing_collections = self.db.list_collection_names()
            if collection_name not in existing_collections:
                raise Exception(f"Collection {collection_name} was not created successfully")
            logger.info(f"✓ Verified collection exists: {collection_name}")
        except Exception as e:
            logger.error(f"Collection verification failed: {e}")
            raise

        # Insert in batches
        logger.info('Inserting documents into AstraDB...')
        batch_size = 20
        for i in range(0, len(documents), batch_size):
            batch = documents[i:i + batch_size]
            try:
                result = collection.insert_many(batch)
                logger.info(f"  Inserted batch {i // batch_size + 1}/{(len(documents) + batch_size - 1) // batch_size}")
            except Exception as e:
                logger.error(f"Error inserting batch {i // batch_size + 1}: {e}")
                logger.error(f"Error type: {type(e).__name__}")
                raise
        
        baseline_id = f"baseline_{int(datetime.now().timestamp())}"

        logger.info(f"{'='*60}")
        logger.info('✓ Baseline ingestion complete')
        logger.info(f"Baseline ID: {baseline_id}")
        logger.info(f"Questions ingested: {len(documents)}")
        logger.info(f"{'='*60}\n")

        return {
            'success': True,
            'baseline_id': baseline_id,
            'questions_ingested': len(documents),
            'collection_name': collection_name,
            'description': description
        }
    
    def verify_equivalence(self, question1, question2):
        """Verify if two questions are semantically equivalent using LLM"""
        if not ENABLE_LLM_VERIFICATION:
            return False
        
        try:
            model = self.initialize_watsonx_model()
            if model is None:
                return False
            
            prompt = f"""Are these two questions asking for the same information? Answer ONLY with "YES" or "NO".

Question A: {question1}
Question B: {question2}

Answer:"""
            
            messages = [{'role': 'user', 'content': prompt}]
            response = model.chat(messages=messages)
            answer = response['choices'][0]['message']['content'].strip().upper()
            
            return answer.startswith('YES')
        except Exception as e:
            logger.error(f"Error in LLM verification: {e}")
            return False
    
    def get_baseline_context(self, question, top_k=5):
        """
        Get relevant Q&A pairs from baseline for RAG context
        
        Args:
            question: The current question to answer
            top_k: Number of top matches to return
            
        Returns:
            Formatted context string with baseline examples
        """
        try:
            collection_name = "delta_baseline"
            collection = self.db.get_collection(collection_name)
            
            # Generate embedding for the question
            embedding = self.generate_embedding(question)
            
            # Query for similar questions
            results = list(collection.find(
                {},
                sort={'$vector': embedding},
                limit=top_k,
                include_similarity=True
            ))
            
            if not results:
                return "No baseline examples available."
            
            context = ""
            for idx, result in enumerate(results, 1):
                similarity = result.get('$similarity', 0)
                q = result.get('question_text', '')
                a = result.get('answer_text', '')
                
                # Skip if answer is marked as unanswered
                if a == 'UNANSWERED':
                    continue
                
                context += f"Baseline Example {idx} (similarity: {similarity:.2%}):\n"
                context += f"Q: {q}\n"
                context += f"A: {a}\n\n"
            
            if not context:
                context = "No answered baseline examples found."
            
            return context.strip()
            
        except Exception as e:
            logger.error(f"Error getting baseline context: {e}")
            return "Error retrieving baseline context."
    
    def answer_with_llm(self, question, baseline_context):
        """
        Use LLM to generate answer based on baseline context
        
        Args:
            question: The question to answer
            baseline_context: Context from baseline Q&A pairs
            
        Returns:
            Generated answer
        """
        try:
            if self.watsonx_model is None:
                self.watsonx_model = self.initialize_watsonx_model()
            
            messages = [
                {
                    "role": "system",
                    "content": """You are an intelligent questionnaire completion assistant. Your role is to answer questions for the CURRENT year's questionnaire by adapting answers from PREVIOUS year's baseline questionnaire.

CRITICAL INSTRUCTIONS:
1. You are answering questions for a NEW questionnaire based on PREVIOUS answers
2. Use the baseline examples as reference, but adapt them if needed for the current question
3. If a question depends on other questions (e.g., "If yes to Q5, explain..."), respond with: "Dependent on previous answer - requires manual review"
4. If the baseline shows the question was unanswered, respond with: "Not answered in baseline - requires new input"
5. Be concise and professional - match the style of baseline answers
6. Do NOT mention that you're using baseline data or reference materials
7. Do NOT include meta-commentary
8. If baseline examples are not relevant, respond with: "No relevant baseline data - requires manual completion"

FORMATTING:
- For yes/no questions: Answer "Yes" or "No" followed by brief details if needed
- For descriptive questions: Provide 1-3 sentences maximum
- Match the tone and format of baseline examples
- Maintain professional business language

Remember: You are filling out the current year's questionnaire using last year's answers as guidance."""
                },
                {
                    "role": "user",
                    "content": f"""Using the following baseline examples from the previous year's questionnaire, answer the current question.

BASELINE EXAMPLES FROM PREVIOUS YEAR:
{baseline_context}

CURRENT QUESTION TO ANSWER:
{question}

Provide your answer:"""
                }
            ]
            
            response = self.watsonx_model.chat(messages=messages)
            answer = response["choices"][0]["message"]["content"]
            
            if not answer or answer.strip() == "":
                return "Error: Empty response from model"
            
            return answer
            
        except Exception as e:
            logger.error(f"Error generating LLM answer: {e}")
            return f"Error generating answer: {str(e)}"
    
    def process_delta(self, current_file_path, threshold=None, use_llm=None):
        """
        Process current questionnaire against baseline
        
        Args:
            current_file_path: Path to current year questionnaire
            threshold: Similarity threshold (optional)
            use_llm: Whether to use LLM generation (None = use env var, True/False = override)
        """
        if not self.is_available():
            raise Exception('Delta Service not available')

        effective_threshold = threshold if threshold is not None else MEDIUM_CONFIDENCE_THRESHOLD
        use_llm_mode = use_llm if use_llm is not None else USE_LLM_GENERATION

        logger.info(f"\n{'='*60}")
        logger.info(f'DELTA SERVICE - Delta Processing ({"LLM" if use_llm_mode else "Copy/Paste"} Mode)')
        logger.info(f"{'='*60}")
        logger.info(f"Current file: {current_file_path}")
        if use_llm_mode:
            logger.info(f"Using LLM generation with baseline RAG context")
        else:
            logger.info(f"Using copy/paste with similarity threshold: {effective_threshold}")

        # Extract questions using smart column detection
        current_questions = self.extract_questions(current_file_path)

        logger.info(f"Extracted {len(current_questions)} questions from current file")

        # Load workbook for updating cells later
        workbook = openpyxl.load_workbook(current_file_path, data_only=False)

        # Get baseline collection
        collection_name = "delta_baseline"
        collection = self.db.get_collection(collection_name)

        # Process each question
        matches = []
        unmatched = []
        high_confidence_count = 0
        medium_confidence_count = 0
        low_confidence_count = 0
        
        logger.info(f'\nProcessing questions...')
        
        for i, current in enumerate(current_questions):
            # Skip if already has an answer
            answer_str = current['current_answer']
            unanswered_values = [
                'unanswered',
                'tbd', 'to be determined',
                'pending',
                'todo',
                'blank',
                'empty'
            ]

            answer_lower = answer_str.lower().strip() if answer_str else ''
            is_answered = (
                answer_str and  # Not empty
                not answer_str.isspace() and  # Not just whitespace
                answer_lower not in unanswered_values and  # Explicitly unanswered values
                not (len(answer_str) == 1 and answer_str in ['-', '_']) and  # Not single dash/underscore
                answer_str not in ['--', '---']  # Not multiple dashes
            )

            if is_answered:
                logger.info(f"  [{i + 1}/{len(current_questions)}] Skipping (already answered): {current['question'][:60]}...")
                continue
            
            if use_llm_mode:
                # LLM MODE: Generate answer using baseline context
                try:
                    baseline_context = self.get_baseline_context(current['question'], top_k=5)
                    generated_answer = self.answer_with_llm(current['question'], baseline_context)
                    
                    # Update Excel
                    self.update_excel_cell(workbook, current['sheet'], current['row'],
                                         current['answer_col'], generated_answer)
                    
                    matches.append({
                        'current_question': current['question'],
                        'current_sheet': current['sheet'],
                        'current_row': current['row'],
                        'matched_question': 'LLM Generated',
                        'similarity_score': 1.0,
                        'confidence': 'LLM',
                        'answer_reused': generated_answer[:100] + '...' if len(generated_answer) > 100 else generated_answer,
                        'baseline_sheet': 'Multiple',
                        'baseline_row': 0,
                        'llm_verified': True
                    })
                    
                    high_confidence_count += 1
                    logger.info(f"  [{i + 1}/{len(current_questions)}] LLM generated: {current['question'][:60]}...")
                    
                except Exception as e:
                    logger.error(f"  [{i + 1}/{len(current_questions)}] LLM error: {str(e)}")
                    unmatched.append({
                        'question': current['question'],
                        'sheet': current['sheet'],
                        'row': current['row'],
                        'reason': f'LLM generation failed: {str(e)}'
                    })
                    
            else:
                # COPY/PASTE MODE: Find similar question and copy answer
                # Generate embedding
                embedding = self.generate_embedding(current['question'])
                
                # Query for similar questions
                results = list(collection.find(
                    {},
                    sort={'$vector': embedding},
                    limit=3,
                    include_similarity=True
                ))
                
                if len(results) == 0:
                    unmatched.append({
                        'question': current['question'],
                        'sheet': current['sheet'],
                        'row': current['row'],
                        'reason': 'No baseline data available'
                    })
                    logger.info(f"  [{i + 1}/{len(current_questions)}] No match: {current['question'][:60]}...")
                    continue
                
                top_match = results[0]
                similarity = top_match.get('$similarity', 0)
                
                if similarity >= HIGH_CONFIDENCE_THRESHOLD:
                    # High confidence - auto-populate
                    matches.append({
                        'current_question': current['question'],
                        'current_sheet': current['sheet'],
                        'current_row': current['row'],
                        'matched_question': top_match['question_text'],
                        'similarity_score': similarity,
                        'confidence': 'HIGH',
                        'answer_reused': top_match['answer_text'],
                        'baseline_sheet': top_match['metadata']['sheet_name'],
                        'baseline_row': top_match['metadata']['row_number'],
                        'llm_verified': False
                    })
                    
                    # Update Excel
                    self.update_excel_cell(workbook, current['sheet'], current['row'],
                                         current['answer_col'], top_match['answer_text'])
                    high_confidence_count += 1
                    
                    logger.info(f"  [{i + 1}/{len(current_questions)}] HIGH match ({similarity:.3f}): {current['question'][:60]}...")
                
                elif similarity >= effective_threshold:
                    # Medium confidence - verify with LLM
                    verified = self.verify_equivalence(current['question'], top_match['question_text'])
                    
                    if verified:
                        matches.append({
                            'current_question': current['question'],
                            'current_sheet': current['sheet'],
                            'current_row': current['row'],
                            'matched_question': top_match['question_text'],
                            'similarity_score': similarity,
                            'confidence': 'MEDIUM',
                            'answer_reused': top_match['answer_text'],
                            'baseline_sheet': top_match['metadata']['sheet_name'],
                            'baseline_row': top_match['metadata']['row_number'],
                            'llm_verified': True
                        })
                        
                        self.update_excel_cell(workbook, current['sheet'], current['row'],
                                             current['answer_col'], top_match['answer_text'])
                        medium_confidence_count += 1
                        
                        logger.info(f"  [{i + 1}/{len(current_questions)}] MEDIUM match ({similarity:.3f}, LLM verified): {current['question'][:60]}...")
                    else:
                        unmatched.append({
                            'question': current['question'],
                            'sheet': current['sheet'],
                            'row': current['row'],
                            'reason': 'LLM verification failed',
                            'best_similarity': similarity,
                            'best_match': top_match['question_text']
                        })
                        
                        logger.info(f"  [{i + 1}/{len(current_questions)}] No match (LLM rejected): {current['question'][:60]}...")
                
                else:
                    # Low confidence - leave blank
                    unmatched.append({
                        'question': current['question'],
                        'sheet': current['sheet'],
                        'row': current['row'],
                        'reason': 'Similarity below threshold',
                        'best_similarity': similarity,
                        'best_match': top_match['question_text']
                    })
                    low_confidence_count += 1
                    
                    logger.info(f"  [{i + 1}/{len(current_questions)}] LOW match ({similarity:.3f}): {current['question'][:60]}...")
        
        # Save processed workbook
        output_filename = f"delta-processed-{int(datetime.now().timestamp())}.xlsx"
        output_path = Path(current_file_path).parent / output_filename
        
        workbook.save(output_path)
        
        logger.info(f"{'='*60}")
        logger.info('✓ Delta processing complete')
        logger.info(f"Total questions: {len(current_questions)}")
        logger.info(f"Auto-answered: {len(matches)} ({len(matches) / len(current_questions) * 100:.1f}%)")
        logger.info(f"  - High confidence: {high_confidence_count}")
        logger.info(f"  - Medium confidence: {medium_confidence_count}")
        logger.info(f"Left blank: {len(unmatched)}")
        logger.info(f"Output file: {output_path}")
        logger.info(f"{'='*60}\n")
        
        return {
            'success': True,
            'processing_summary': {
                'total_questions': len(current_questions),
                'auto_answered': len(matches),
                'left_blank': len(unmatched),
                'high_confidence': high_confidence_count,
                'medium_confidence': medium_confidence_count,
                'low_confidence': len(unmatched),
                'completion_rate': f"{len(matches) / len(current_questions) * 100:.1f}%"
            },
            'matches': matches,
            'unmatched': unmatched,
            'output_path': str(output_path),
            'output_filename': output_filename
        }
    
    def update_excel_cell(self, workbook, sheet_name, row, col, value):
        """Update a cell in the Excel workbook"""
        worksheet = workbook[sheet_name]
        cell = worksheet.cell(row=row, column=col + 1)  # Excel is 1-indexed
        cell.value = value
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    def get_baseline_info(self):
        """Get information about the current baseline"""
        if not self.is_available():
            return None

        try:
            collection_name = "delta_baseline"
            collections = self.db.list_collection_names()

            if collection_name not in collections:
                return None

            collection = self.db.get_collection(collection_name)

            # Get count and first document for metadata
            try:
                # Count documents
                cursor = collection.find({}, limit=1, projection={"metadata": 1})
                first_doc = next(cursor, None)

                if first_doc:
                    metadata = first_doc.get('metadata', {})
                    return {
                        'exists': True,
                        'collection_name': collection_name,
                        'description': metadata.get('description', ''),
                        'user_email': metadata.get('user_email', ''),
                        'ingested_at': metadata.get('ingested_at', '')
                    }
                else:
                    return {
                        'exists': True,
                        'collection_name': collection_name,
                        'empty': True
                    }
            except Exception as e:
                logger.error(f"Error reading baseline metadata: {e}")
                return {
                    'exists': True,
                    'collection_name': collection_name,
                    'error': str(e)
                }

        except Exception as e:
            logger.error(f"Error getting baseline info: {e}")
            return None

    def clear_baseline(self):
        """Clear the baseline collection"""
        if not self.is_available():
            raise Exception('Delta Service not available')

        collection_name = "delta_baseline"

        try:
            collections = self.db.list_collection_names()
            if collection_name in collections:
                collection = self.db.get_collection(collection_name)
                delete_result = collection.delete_many({})
                logger.info(f"✓ Cleared baseline collection")
                return {'success': True, 'message': 'Baseline cleared'}
            else:
                return {'success': True, 'message': 'No baseline exists'}
        except Exception as e:
            logger.error(f"Error clearing baseline: {e}")
            raise


# Singleton instance
delta_service = DeltaService()

# Made with Bob
