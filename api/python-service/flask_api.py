#!/usr/bin/env python3
"""
Flask API Service - Bridge between Node.js Express and Python document processor
Provides REST API endpoints for the Python-based document processing
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os
import tempfile
import logging
from pathlib import Path
from document_processor import process_document, initialize_model, detect_qa_columns_in_sheet
import pandas as pd
import openpyxl

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Create Flask app
app = Flask(__name__)
CORS(app)

# Cache model to avoid re-initialization
_model_cache = None


def get_model():
    """Get or initialize the model (cached)"""
    global _model_cache
    if _model_cache is None:
        logger.info("Initializing model...")
        _model_cache = initialize_model()
        logger.info("Model initialized")
    return _model_cache


@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        "status": "ok",
        "service": "Python Document Processor",
        "rag_enabled": os.getenv("ASTRA_DB_API_ENDPOINT") is not None
    })


@app.route('/process', methods=['POST'])
def process_excel():
    """
    Process Excel file with RAG-powered AI completion.

    Expects multipart/form-data with:
    - file: Excel file
    - context: Optional context string

    Returns JSON with processing summary, preview data, and a download_filename
    that can be fetched via GET /python/download/<filename>.
    """
    input_path = None

    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400

        file = request.files['file']
        context = request.form.get('context', '')

        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400

        if not file.filename.endswith('.xlsx'):
            return jsonify({"error": "File must be .xlsx format"}), 400

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_input:
            file.save(temp_input.name)
            input_path = temp_input.name

        logger.info(f"Processing file: {file.filename}")

        result = process_document(input_path, context=context)

        if not result.get('success', False):
            error_msg = result.get('error', 'Unknown processing error')
            logger.error(f"Processing failed: {error_msg}")
            return jsonify({
                "error": error_msg,
                "details": result.get('errors', [])
            }), 500

        output_path = result.get('output_path')
        if not output_path or not os.path.exists(output_path):
            return jsonify({"error": "Output file not generated"}), 500

        # Store output file for later download (keyed by download filename)
        output_filename = f"{Path(file.filename).stem}_completed.xlsx"
        if not hasattr(app, 'python_files'):
            app.python_files = {}
        app.python_files[output_filename] = output_path

        return jsonify({
            "success": True,
            "questions_answered": result.get('questions_answered', 0),
            "sheets_processed": result.get('sheets_processed', 0),
            "total_sheets": result.get('total_sheets', 0),
            "details": result.get('details', []),
            "qa_pairs": result.get('qa_pairs', []),
            "download_filename": output_filename
        })

    except Exception as e:
        logger.error(f"Error processing file: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

    finally:
        if input_path and os.path.exists(input_path):
            try:
                os.unlink(input_path)
            except Exception as e:
                logger.warning(f"Failed to clean up input file: {e}")


@app.route('/python/download/<path:filename>', methods=['GET'])
def download_python_file(filename):
    """Download a processed Excel file stored by /process"""
    try:
        if not hasattr(app, 'python_files') or filename not in app.python_files:
            return jsonify({"error": "File not found or expired"}), 404

        filepath = app.python_files[filename]

        if not os.path.exists(filepath):
            del app.python_files[filename]
            return jsonify({"error": "File not found"}), 404

        response = send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        @response.call_on_close
        def cleanup():
            try:
                if os.path.exists(filepath):
                    os.unlink(filepath)
                if hasattr(app, 'python_files') and filename in app.python_files:
                    del app.python_files[filename]
            except Exception as e:
                logger.warning(f"Failed to clean up {filepath}: {e}")

        return response

    except Exception as e:
        logger.error(f"Error downloading python file: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/download/<path:filepath>', methods=['GET'])
def download_file(filepath):
    """Download processed file"""
    try:
        if not os.path.exists(filepath):
            return jsonify({"error": "File not found"}), 404
        
        filename = request.args.get('filename', 'completed.xlsx')
        
        response = send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Clean up file after sending
        @response.call_on_close
        def cleanup():
            try:
                if os.path.exists(filepath):
                    os.unlink(filepath)
                    logger.info(f"Cleaned up temp file: {filepath}")
            except Exception as e:
                logger.warning(f"Failed to clean up {filepath}: {e}")
        
        return response
    
    except Exception as e:
        logger.error(f"Error downloading file: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/detect-columns', methods=['POST'])
def detect_columns():
    """
    Detect Q&A columns in an Excel sheet
    
    Expects multipart/form-data with:
    - file: Excel file
    - sheet_name: Optional sheet name (defaults to first sheet)
    
    Returns JSON with detected column names
    """
    temp_path = None
    
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400
        
        file = request.files['file']
        sheet_name = request.form.get('sheet_name', 0)
        
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        # Save to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            file.save(temp_file.name)
            temp_path = temp_file.name
        
        # Read the sheet
        df = pd.read_excel(temp_path, sheet_name=sheet_name)
        
        if df.empty:
            return jsonify({"error": "Sheet is empty"}), 400
        
        # Detect columns
        model = get_model()
        question_col, answer_col = detect_qa_columns_in_sheet(df, model)
        
        if not question_col or not answer_col:
            return jsonify({"error": "Could not detect Q&A columns"}), 400
        
        return jsonify({
            "success": True,
            "question_column": question_col,
            "answer_column": answer_col
        })
    
    except Exception as e:
        logger.error(f"Error detecting columns: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500
    
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.unlink(temp_path)
            except Exception as e:
                logger.warning(f"Failed to clean up temp file: {e}")


# ============================================================================
# Delta Tool API Endpoints
# ============================================================================

try:
    from delta_service import delta_service
    DELTA_SERVICE_AVAILABLE = True
except ImportError as e:
    logger.warning(f"Delta service not available: {e}")
    DELTA_SERVICE_AVAILABLE = False
    delta_service = None

@app.route('/delta/status', methods=['GET'])
def delta_status():
    """Check Delta Service status"""
    if not DELTA_SERVICE_AVAILABLE or delta_service is None:
        return jsonify({
            "available": False,
            "service": "Delta Tool",
            "message": "Delta service module not loaded"
        })
    
    try:
        return jsonify({
            "available": delta_service.is_available(),
            "service": "Delta Tool",
            "astradb_configured": os.getenv("ASTRA_DB_API_ENDPOINT") is not None,
            "llm_verification_enabled": os.getenv("DELTA_ENABLE_LLM_VERIFICATION", "true").lower() == "true",
            "similarity_threshold": float(os.getenv("DELTA_SIMILARITY_THRESHOLD", "0.85"))
        })
    except Exception as e:
        logger.error(f"Error checking delta status: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/delta/upload-baseline', methods=['POST'])
def upload_baseline():
    """Upload and ingest baseline questionnaire"""
    temp_path = None

    try:
        # Debug logging
        logger.info(f"Upload baseline request received")
        logger.info(f"  Request files: {list(request.files.keys())}")
        logger.info(f"  Request form: {dict(request.form)}")
        logger.info(f"  Content type: {request.content_type}")

        if not DELTA_SERVICE_AVAILABLE or delta_service is None:
            logger.error("Delta service not available or not loaded")
            return jsonify({
                "error": "Delta Service not available",
                "message": "Delta service module not loaded"
            }), 503

        if 'file' not in request.files:
            logger.error(f"No 'file' in request.files. Available keys: {list(request.files.keys())}")
            return jsonify({"error": "No file provided"}), 400

        file = request.files['file']
        description = request.form.get('description', '')
        user_email = request.headers.get('x-forwarded-email', 'dev.user@ibm.com')

        logger.info(f"  File: {file.filename if file else 'None'}")
        logger.info(f"  Description: {description}")
        logger.info(f"  User: {user_email}")

        if file.filename == '':
            logger.error("File filename is empty")
            return jsonify({"error": "No file selected"}), 400

        if not file.filename.endswith('.xlsx'):
            logger.error(f"Invalid file format: {file.filename}")
            return jsonify({"error": "File must be .xlsx format"}), 400
        
        if not delta_service.is_available():
            return jsonify({
                "error": "Delta Service not available",
                "message": "AstraDB credentials not configured"
            }), 503
        
        # Save uploaded file to temp location
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            file.save(temp_file.name)
            temp_path = temp_file.name
        
        logger.info(f"Ingesting baseline: {file.filename}")

        # Ingest baseline (clears existing baseline)
        result = delta_service.ingest_baseline(temp_path, user_email, description)
        
        return jsonify(result)
    
    except Exception as e:
        logger.error(f"Error uploading baseline: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500
    
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.unlink(temp_path)
            except Exception as e:
                logger.warning(f"Failed to clean up temp file: {e}")


@app.route('/delta/process', methods=['POST'])
def process_delta():
    """Process current year questionnaire against baseline"""
    temp_path = None
    
    try:
        if not DELTA_SERVICE_AVAILABLE or delta_service is None:
            return jsonify({
                "error": "Delta Service not available",
                "message": "Delta service module not loaded"
            }), 503
        
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400
        
        file = request.files['file']
        similarity_threshold = request.form.get('similarity_threshold')
        use_llm = request.form.get('use_llm', '').lower() == 'true'
        
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({"error": "File must be .xlsx format"}), 400
        
        if not delta_service.is_available():
            return jsonify({
                "error": "Delta Service not available",
                "message": "AstraDB credentials not configured"
            }), 503
        
        # Save uploaded file to temp location
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            file.save(temp_file.name)
            temp_path = temp_file.name
        
        logger.info(f"Processing delta: {file.filename} (LLM mode: {use_llm})")

        # Process delta
        threshold = float(similarity_threshold) if similarity_threshold else None
        result = delta_service.process_delta(temp_path, threshold, use_llm=use_llm)
        
        # Store output file path in a simple in-memory dict with the filename as key
        # In production, use Redis or a proper cache
        output_filename = result['output_filename']
        if not hasattr(app, 'delta_files'):
            app.delta_files = {}
        app.delta_files[output_filename] = result['output_path']
        
        # Return JSON with metadata and download endpoint
        return jsonify({
            "success": True,
            "processing_summary": result['processing_summary'],
            "matches": result['matches'],
            "unmatched": result['unmatched'],
            "download_filename": output_filename
        })
    
    except Exception as e:
        logger.error(f"Error processing delta: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500
    
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.unlink(temp_path)
            except Exception as e:
                logger.warning(f"Failed to clean up temp file: {e}")


@app.route('/delta/download/<filename>', methods=['GET'])
def download_delta_file(filename):
    """Download processed delta file by filename"""
    try:
        # Get file path from in-memory storage
        if not hasattr(app, 'delta_files') or filename not in app.delta_files:
            return jsonify({"error": "File not found or expired"}), 404
        
        filepath = app.delta_files[filename]
        
        if not os.path.exists(filepath):
            # Clean up stale entry
            del app.delta_files[filename]
            return jsonify({"error": "File not found"}), 404
        
        response = send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Clean up file after sending
        @response.call_on_close
        def cleanup():
            try:
                if os.path.exists(filepath):
                    os.unlink(filepath)
                    logger.info(f"Cleaned up temp file: {filepath}")
                # Remove from cache
                if hasattr(app, 'delta_files') and filename in app.delta_files:
                    del app.delta_files[filename]
            except Exception as e:
                logger.warning(f"Failed to clean up {filepath}: {e}")
        
        return response
    
    except Exception as e:
        logger.error(f"Error downloading delta file: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/delta/baseline-info', methods=['GET'])
def get_baseline_info():
    """Get information about the current baseline"""
    try:
        if not DELTA_SERVICE_AVAILABLE or delta_service is None:
            return jsonify({
                "available": False,
                "baseline": None,
                "message": "Delta Service not configured"
            })

        if not delta_service.is_available():
            return jsonify({
                "available": False,
                "baseline": None,
                "message": "Delta Service not configured"
            })

        baseline_info = delta_service.get_baseline_info()

        return jsonify({
            "available": True,
            "baseline": baseline_info
        })

    except Exception as e:
        logger.error(f"Error getting baseline info: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/delta/preview/<path:filename>', methods=['GET'])
def preview_delta_file(filename):
    """Get sheet preview data from a processed delta file"""
    try:
        if not hasattr(app, 'delta_files') or filename not in app.delta_files:
            return jsonify({"error": "File not found or expired"}), 404

        filepath = app.delta_files[filename]

        if not os.path.exists(filepath):
            if hasattr(app, 'delta_files') and filename in app.delta_files:
                del app.delta_files[filename]
            return jsonify({"error": "File not found"}), 404

        workbook = openpyxl.load_workbook(filepath, data_only=True)

        sheets = []
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 51:  # header + 50 data rows
                    break
                row_data = [str(cell) if cell is not None else '' for cell in row]
                if any(cell for cell in row_data):
                    rows.append(row_data)

            if rows:
                sheets.append({
                    'name': sheet_name,
                    'headers': rows[0] if rows else [],
                    'rows': rows[1:] if len(rows) > 1 else []
                })

        return jsonify({'success': True, 'sheets': sheets})

    except Exception as e:
        logger.error(f"Error previewing delta file: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/delta/baseline', methods=['DELETE'])
def clear_baseline():
    """Clear the baseline collection"""
    try:
        if not DELTA_SERVICE_AVAILABLE or delta_service is None:
            return jsonify({"error": "Delta Service not available"}), 503

        if not delta_service.is_available():
            return jsonify({"error": "Delta Service not available"}), 503

        result = delta_service.clear_baseline()
        return jsonify(result)

    except Exception as e:
        logger.error(f"Error clearing baseline: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


# ============================================================================
# Application Entry Point
# ============================================================================

if __name__ == '__main__':
    port = int(os.getenv('PYTHON_SERVICE_PORT', 5000))
    logger.info(f"Starting Python Document Processor API on port {port}")
    logger.info(f"Delta Service Available: {DELTA_SERVICE_AVAILABLE}")
    app.run(host='0.0.0.0', port=port, debug=False)

# Made with Bob

